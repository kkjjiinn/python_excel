import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment

window = tk.Tk()
window.title("음식점 관리 프로그램")
window.geometry("600x600")

# 음식점 정보를 저장할 리스트
restaurant_data = []

# 제목 셀 병합
def merge_title(ws):
    ws.merge_cells('A1:H2')
    title_cell = ws['A1']
    title_cell.value = '음식점'
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

# 항목명 작성
def write_header(ws):
    header = ["이름", "주소", "거리", "메뉴", "별점", "리뷰", "특징", "전화번호"]
    ws.append([])
    ws.append(header)
    ws.row_dimensions[3].height = 20
    for col in ws.iter_cols(min_col=1, max_col=8, min_row=3, max_row=3):
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')

# 내용 작성
def write_data(ws):
    for restaurant in restaurant_data:
        ws.append(restaurant)

# 저장 버튼 클릭 시 실행될 함수
def save_data():
    name = entry_name.get()
    address = entry_address.get()
    distance = entry_distance.get()
    menu = entry_menu.get()
    rating = entry_rating.get()
    review = entry_review.get()
    features = entry_features.get()
    phone = entry_phone.get()

    restaurant = [name, address, distance, menu, rating, review, features, phone]
    restaurant_data.append(restaurant)

    messagebox.showinfo("성공", "음식점 정보가 저장되었습니다.")
    clear_entries()

# 삭제 버튼 클릭 시 실행될 함수
def delete_data():
    if messagebox.askyesno("삭제", "정말로 모든 음식점 정보를 삭제하시겠습니까?"):
        restaurant_data.clear()
        messagebox.showinfo("성공", "모든 음식점 정보가 삭제되었습니다.")
        clear_entries()

# 엔트리 값들 초기화
def clear_entries():
    entry_name.delete(0, tk.END)
    entry_address.delete(0, tk.END)
    entry_distance.delete(0, tk.END)
    entry_menu.delete(0, tk.END)
    entry_rating.delete(0, tk.END)
    entry_review.delete(0, tk.END)
    entry_features.delete(0, tk.END)
    entry_phone.delete(0, tk.END)

# GUI 내용 확인 및 엑셀 내보내기
def view_data():
    wb = Workbook()
    ws = wb.active

    merge_title(ws)
    write_header(ws)
    write_data(ws)

    messagebox.showinfo("성공", "음식점 정보를 엑셀 파일로 내보냈습니다.")

    # 엑셀 파일 저장
    wb.save("restaurants.xlsx")

# 이름
label_name = tk.Label(window, text="이름:")
label_name.pack()
entry_name = tk.Entry(window)
entry_name.pack()

# 주소
label_address = tk.Label(window, text="주소:")
label_address.pack()
entry_address = tk.Entry(window)
entry_address.pack()

# 거리
label_distance = tk.Label(window, text="거리:")
label_distance.pack()
entry_distance = tk.Entry(window)
entry_distance.pack()

# 메뉴
label_menu = tk.Label(window, text="메뉴:")
label_menu.pack()
entry_menu = tk.Entry(window)
entry_menu.pack()

# 별점
label_rating = tk.Label(window, text="별점:")
label_rating.pack()
entry_rating = tk.Entry(window)
entry_rating.pack()

# 리뷰
label_review = tk.Label(window, text="리뷰:")
label_review.pack()
entry_review = tk.Entry(window)
entry_review.pack()

# 특징
label_features = tk.Label(window, text="특징:")
label_features.pack()
entry_features = tk.Entry(window)
entry_features.pack()

# 전화번호
label_phone = tk.Label(window, text="전화번호:")
label_phone.pack()
entry_phone = tk.Entry(window)
entry_phone.pack()

# 저장 버튼
save_button = tk.Button(window, text="저장", command=save_data)
save_button.pack()

# 삭제 버튼
delete_button = tk.Button(window, text="삭제", command=delete_data)
delete_button.pack()

# GUI 내용 확인 및 엑셀 내보내기 버튼
view_button = tk.Button(window, text="GUI 내용 확인 및 엑셀 내보내기", command=view_data)
view_button.pack()

window.mainloop()
